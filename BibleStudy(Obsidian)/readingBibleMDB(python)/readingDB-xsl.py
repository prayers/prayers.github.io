import sqlite3
import openpyxl

# 连接数据库
conn = sqlite3.connect('bible_简体中文和合本.db')
cursor = conn.cursor()

# 创建 Excel 工作簿
workbook = openpyxl.Workbook()
sheet = workbook.active

# 写入表头
sheet['A1'] = '经卷'
sheet['B1'] = '章节'
sheet['C1'] = '节数'
sheet['D1'] = '经文内容'

# 查询所有经文信息
sql = """SELECT VolumeSN, ChapterSN, VerseSN, Lection FROM Bible"""
cursor.execute(sql)
results = cursor.fetchall()

# 遍历结果并写入 Excel
row = 2
for result in results:
    sheet.cell(row=row, column=1, value=result[0])  # 经卷
    sheet.cell(row=row, column=2, value=result[1])  # 章节
    sheet.cell(row=row, column=3, value=result[2])  # 节数
    sheet.cell(row=row, column=4, value=result[3])  # 经文内容
    row += 1

# 保存 Excel 文件
workbook.save('bible_data.xlsx')

# 关闭数据库连接
conn.close()