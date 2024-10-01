import sqlite3
import openpyxl

def export_bible_to_excel(db_file, table_name, columns_in_db, output_file, columns):
    """
    将 SQLite 数据库中的圣经经文数据导出到 Excel 文件

    Args:
        db_file: 数据库文件名
        table_name: 表名
        output_file: 输出 Excel 文件名
        columns: 字段列表，对应 Excel 的列名
    """

    try:
        # 连接数据库
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()

        # 创建 Excel 工作簿
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # 写入表头
        for i, col in enumerate(columns, start=1):
            sheet.cell(row=1, column=i, value=col)

        # 查询所有经文信息
        sql = f"SELECT {', '.join(columns_in_db)} FROM {table_name}"
        cursor.execute(sql)

        # 遍历结果并写入 Excel
        row = 2
        for result in cursor:
            for i, value in enumerate(result, start=1):
                sheet.cell(row=row, column=i, value=value)
            row += 1

        # 保存 Excel 文件
        workbook.save(output_file)

    except Exception as e:
        print(f"导出数据失败: {e}")
    finally:
        # 关闭数据库连接
        if conn:
            conn.close()

# 使用示例
db_file = 'bible_简体中文和合本.db'
table_name = 'Bible'
output_file = 'bible_data.xlsx'
columns_in_db = ['VolumeSN', 'ChapterSN', 'VerseSN', 'Lection']
columns = ['经卷', '章节', '节数', '经文内容']  # 这里使用统一的中文字段名

export_bible_to_excel(db_file, table_name, columns_in_db, output_file, columns)